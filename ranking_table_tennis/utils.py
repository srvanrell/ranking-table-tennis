import os
from typing import List, Tuple, Any

from ranking_table_tennis import models
from ranking_table_tennis.models import cfg
from openpyxl.styles import Font, Alignment
import gspread
from gspread.utils import rowcol_to_a1
import pandas as pd
import pickle

from gspread_dataframe import set_with_dataframe

__author__ = 'sebastian'


def get_tournament_sheet_names_ordered() -> List[str]:
    tournaments_xlsx = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    filter_key = cfg["sheetname"]["tournaments_key"]
    df_tournaments = pd.read_excel(tournaments_xlsx, sheet_name=None, header=None)
    sheet_names = [s for s in df_tournaments.keys() if filter_key in s]

    return sorted(sheet_names)


def _bold_and_center(ws, to_bold: List[str], to_center: List[str]) -> None:
    """Bold and center cells in given worksheet (ws)"""
    for colrow in to_bold:
        cell = ws[colrow]
        cell.font = Font(bold=True)
    for colrow in to_center:
        cell = ws[colrow]
        cell.alignment = Alignment(horizontal='center')


def save_ranking_sheet(tid: str, tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                       upload: bool = False) -> None:
    if tid == cfg["aux"]["initial tid"]:
        sheet_name = cfg["sheetname"]["initial_ranking"]
        xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    else:
        xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["rankings_filename"]
        sheet_name = tournaments[tid].iloc[0].sheet_name
        sheet_name = sheet_name.replace(cfg["sheetname"]["tournaments_key"],
                                        cfg["sheetname"]["rankings_key"])

    sorted_rankings_df = rankings.ranking_df.sort_values("rating", ascending=False)
    sorted_rankings_df.loc[:, "active"] = sorted_rankings_df.loc[:, "active"].apply(lambda x:  cfg["activeplayer"][x])
    sorted_rankings_df.insert(2, "name", sorted_rankings_df.loc[:, "pid"].apply(lambda pid: players[pid]["name"]))
    sorted_rankings_df.loc[:, "date"] = sorted_rankings_df.loc[:, "date"].apply(lambda d: d.strftime("%Y %m %d"))

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["tid", "Tournament name", "Date", "Location",
                                                  "pid", "Player", "Rating", "Category", "Active Player"]]
        columns = ["tid", "tournament_name", "date", "location", "pid", "name", "rating", "category", "active"]
        sorted_rankings_df.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=columns)

    if upload:
        upload_sheet_from_df(cfg["io"]["tournaments_spreadsheet_id"], cfg["sheetname"]["initial_ranking"],
                             sorted_rankings_df.loc[:, columns], headers)


def save_players_sheet(players: models.Players, upload=False) -> None:
    sorted_players_df = players.players_df.sort_values("name")
    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    sheet_name = cfg["sheetname"]["players"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["Player", "Association", "City"]]
        sorted_players_df.to_excel(writer, sheet_name=sheet_name, index_label=cfg["labels"]["pid"], header=headers)

    if upload:
        headers_df = [cfg["labels"]["pid"]] + headers
        upload_sheet_from_df(cfg["io"]["tournaments_spreadsheet_id"], cfg["sheetname"]["players"],
                             sorted_players_df, headers_df, include_index=True)  # index is pid


def _get_ws_from_spreadsheet(sheet_name: str, spreadsheet_id: str):
    gc = _get_gc()
    wb = gc.open_by_key(spreadsheet_id)
    if sheet_name not in [ws.title for ws in wb.worksheets()]:
        wb.add_worksheet(sheet_name, rows=1, cols=1)
    ws = wb.worksheet(sheet_name)

    return ws


def upload_sheet_from_df(spreadsheet_id: str, sheet_name: str, df: pd.DataFrame, headers: List[str] = None,
                         include_index: bool = False, include_df_headers: bool = True) -> None:
    """
    Saves headers and df data into given sheet_name in spreadsheet_id.
    If sheet_name does not exist, it will be created.

    :param include_index: will upload DataFrame index as the first column. It is False by default.
    :param include_df_headers: will upload DataFrame column names as the first row. It is True by default.
    :param headers: This list will replace df column names and must have the same length.
    If headers are given, include_df_headers is turn to True.
    """
    print("<<<Saving", sheet_name, "in", spreadsheet_id, sep="\t")
    try:
        worksheet = _get_ws_from_spreadsheet(sheet_name, spreadsheet_id)

        df_headers = df.copy()
        if include_index:
            df_headers.reset_index(inplace=True)
        if headers:
            df_headers.columns = headers
            include_df_headers = True

        set_with_dataframe(worksheet, df_headers, resize=True, #include_index=include_index,
                           include_column_header=include_df_headers)

    except ConnectionError:
        print("<<<FAILED to upload", sheet_name, "in", spreadsheet_id, sep="\t")


def load_players_sheet() -> models.Players:
    tournaments_xlsx = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    players_df = pd.read_excel(tournaments_xlsx, sheet_name=cfg["sheetname"]["players"], header=0)
    print("> Reading", cfg["sheetname"]["players"], "from", tournaments_xlsx, sep="\t")

    players_df.rename(columns={cfg["labels"]["pid"]: "pid",
                               cfg["labels"]["Player"]: "name",
                               cfg["labels"]["Association"]: "affiliation",
                               cfg["labels"]["City"]: "city"},
                      inplace=True)

    players = models.Players(players_df)

    return players


def load_initial_ranking_sheet() -> models.Rankings:
    tournaments_xlsx = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    initial_ranking_df = pd.read_excel(tournaments_xlsx, sheet_name=cfg["sheetname"]["initial_ranking"], header=0)
    print("> Reading", cfg["sheetname"]["initial_ranking"], "from", tournaments_xlsx, sep="\t")

    columns_translations = {cfg["labels"]["tid"]: "tid", cfg["labels"]["Tournament name"]: "tournament_name",
                            cfg["labels"]["Date"]: "date", cfg["labels"]["Location"]: "location",
                            cfg["labels"]["pid"]: "pid", cfg["labels"]["Player"]: "name",
                            cfg["labels"]["Rating"]: "rating", cfg["labels"]["Category"]: "category",
                            cfg["labels"]["Active Player"]: "active"}

    initial_ranking_df.rename(columns=columns_translations, inplace=True)
    initial_ranking_df.loc[:, "active"] = initial_ranking_df.loc[:, "active"].apply(lambda x:
                                                                                    x == cfg["activeplayer"][True])
    initial_ranking = models.Rankings(initial_ranking_df)

    return initial_ranking


def load_tournaments_sheets() -> models.Tournaments:
    tournaments_xlsx = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    filter_key = cfg["sheetname"]["tournaments_key"]
    raw_tournaments = pd.read_excel(tournaments_xlsx, sheet_name=None, header=None)
    sheet_names = sorted([s for s in raw_tournaments.keys() if filter_key in s])

    to_concat = []
    for sheet_name in sheet_names:
        print("> Reading", sheet_name, "from", tournaments_xlsx, sep="\t")
        raw_tournament = raw_tournaments[sheet_name]

        tournament_df = raw_tournament.iloc[5:].copy()
        tournament_df.rename(columns={0: "player_a", 1: "player_b", 2: "sets_a", 3: "sets_b",
                                      4: "round", 5: "category"}, inplace=True)
        tournament_df.insert(0, "location", raw_tournament.iat[2, 1])
        tournament_df.insert(0, "date", raw_tournament.iat[1, 1])
        tournament_df.insert(0, "tournament_name", raw_tournament.iat[0, 1])
        tournament_df.insert(0, "sheet_name", sheet_name)

        to_concat.append(tournament_df)

    tournaments_df = pd.concat(to_concat, ignore_index=True)
    tournaments = models.Tournaments(tournaments_df)

    return tournaments


def _format_diff(new_value: float, prev_value: float) -> str:
    """Return a formated str that indicates +##, -## or ="""
    diff = new_value - prev_value
    diff_str = f"{diff:.0f}"
    if diff > 0:
        diff_str = "+" + diff_str
    elif diff == 0:
        diff_str = "="
    formatted_str = f"{new_value:.0f} ({diff_str})"

    return formatted_str


def _publish_tournament_metadata(ws, tournament_tid: pd.DataFrame) -> None:
    ws.insert_rows(0, 3)
    ws["A1"] = cfg["labels"]["Tournament name"]
    ws["B1"] = tournament_tid["tournament_name"].iloc[0]
    ws.merge_cells('B1:E1')
    ws["A2"] = cfg["labels"]["Date"]
    ws["B2"] = tournament_tid["date"].iloc[0].strftime("%Y %m %d")
    ws.merge_cells('B2:E2')
    ws["A3"] = cfg["labels"]["Location"]
    ws["B3"] = tournament_tid["location"].iloc[0]
    ws.merge_cells('B3:E3')


def _get_writer_mode(xlsx_file_path: str) -> str:
    mode = 'a'
    if not os.path.exists(xlsx_file_path):
        mode = 'w'

    return mode


def _get_writer(xlsx_filename: str, sheet_name: str) -> pd.ExcelWriter:
    writer = pd.ExcelWriter(xlsx_filename, engine='openpyxl', mode=_get_writer_mode(xlsx_filename))
    if sheet_name in writer.book.sheetnames:
        writer.book.remove(writer.book[sheet_name])
    print("<<<Saving", sheet_name, "in", xlsx_filename, sep="\t")

    return writer


def publish_rating_sheet(tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                         tid: str, prev_tid: str, upload=False) -> None:
    """ Format a ranking to be published into a rating sheet
    """
    sheet_name = tournaments[tid]["sheet_name"].iloc[0]
    sheet_name = sheet_name.replace(cfg["sheetname"]["tournaments_key"], cfg["labels"]["Rating"])

    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)

    # Rankings sorted by rating
    this_ranking_df = rankings[tid].sort_values("rating", ascending=False)
    prev_ranking_df = rankings[prev_tid]

    # Filter inactive players or players that didn't played any tournament
    nonzero_points = this_ranking_df.loc[:, rankings.cum_points_cat_columns()].any(axis="columns")
    this_ranking_df = this_ranking_df.loc[this_ranking_df.active | nonzero_points]
    # FIXME there must be a special treatment for fan category

    # Format data and columns to write into the file
    this_ranking_df = this_ranking_df.merge(players.players_df.loc[:, ["name", "city", "affiliation"]], on="pid")
    this_ranking_df = this_ranking_df.merge(prev_ranking_df.loc[:, ["pid", "rating"]], on="pid", suffixes=("", "_prev"))
    this_ranking_df.insert(6, "formatted rating", this_ranking_df.apply(
        lambda row: _format_diff(row['rating'], row['rating_prev']), axis="columns"))

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4", "E4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["Category", "Rating", "Player", "City", "Association"]]
        columns = ["category", "formatted rating", "name", "city", "affiliation"]
        this_ranking_df.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=columns)

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def publish_initial_rating_sheet(tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                                 tid: str, upload=False) -> None:
    """ Format a ranking to be published into a rating sheet
    """
    sheet_name = cfg["sheetname"]["initial_ranking"]
    sheet_name = sheet_name.replace(cfg["sheetname"]["rankings_key"], cfg["labels"]["Rating"])

    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)

    # Rankings sorted by rating
    initial_tid = cfg["aux"]["initial tid"]
    this_ranking_df = rankings[initial_tid].sort_values("rating", ascending=False)

    # Format data and columns to write into the file
    this_ranking_df = this_ranking_df.merge(players.players_df.loc[:, ["name", "city", "affiliation"]], on="pid")

    to_bold = ["A1", "B1", "C1", "D1"]
    to_center = to_bold

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["Rating", "Player", "City", "Association"]]
        columns = ["rating", "name", "city", "affiliation"]
        this_ranking_df.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=columns)

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def save_raw_ranking(rankings: models.Rankings, players: models.Players,
                     tid: str) -> None:
    """Add players name to raw ranking to be save into a spreadsheet."""
    xlsx_filename = cfg["io"]["data_folder"] + f"raw_ranking_{tid}.xlsx"
    print("<<<Saving raw ranking in", xlsx_filename, sep="\t")

    # Rankings sorted by rating
    this_ranking_df = rankings[tid].sort_values("rating", ascending=False)

    # Format data and columns to write into the file
    this_ranking_df = this_ranking_df.merge(players.players_df.loc[:, ["name"]], on="pid")
    this_ranking_df.to_excel(xlsx_filename)


def _keep_name_new_row(df: pd.DataFrame) -> pd.DataFrame:
    """Function to insert row in the dataframe"""
    empty_row = pd.DataFrame({'tid': '', 'pid': '', 'category': '', 'best_round': '',
                              'name': df.loc[df.first_valid_index(), 'name']}, index=[-1])
    df.loc[df.first_valid_index(), 'name'] = ""

    return pd.concat([empty_row, df])


def publish_histories_sheet(tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                            tid: str, prev_tid: str, upload=False) -> None:
    """ Format histories to be published into a sheet"""
    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)
    sheet_name = cfg["sheetname"]["histories"]

    history_df = players.history_df.copy()

    # Match pid to get player's metadata into new columns of history_df
    history_df = pd.merge(history_df, players.players_df, on="pid")
    # Sort histories by name, category and tid
    history_df = history_df.sort_values(["name", "category", "tid"], ascending=[True, False, True])
    # Remove repeated strings to show a cleaner sheet
    history_df.loc[history_df['name'] == history_df['name'].shift(1), "name"] = ""
    # insert an empty row into history. This is a format workaround
    history_df = history_df.groupby('pid', sort=False, group_keys=False).apply(_keep_name_new_row)
    # Remove repeated strings to show a cleaner sheet
    history_df.loc[history_df['category'] == history_df['category'].shift(1), "category"] = ""

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["Player", "Category", "Best Round", "Tournament"]]
        columns = ["name", "category", "best_round", "tid"]
        history_df.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=columns)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def publish_rating_details_sheet(tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                                 tid: str, prev_tid: str, upload) -> None:
    """Format and publish rating details of given tournament into a sheet"""

    sheet_name = tournaments[tid]["sheet_name"].iloc[0]
    sheet_name = sheet_name.replace(cfg["sheetname"]["tournaments_key"], cfg["sheetname"]["rating_details_key"])

    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)

    details = rankings.get_rating_details(tid)

    details["winner_name_rating"] = details['winner'] + " (" + details['winner_rating'].astype(int).astype(str) + ")"
    details["loser_name_rating"] = details['loser'] + " (" + details['loser_rating'].astype(int).astype(str) + ")"
    details["diff_rating"] = (details['winner_rating'] - details['loser_rating']).astype(int).astype(str)
    details["factor"] /= cfg["aux"]["rating factor"]

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4", "E4", "F4", "G4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["Winner", "Loser", "Difference", "Winner Points", "Loser Points",
                                                  "Round", "Category", "Factor"]]
        columns = ["winner_name_rating", "loser_name_rating", "diff_rating", "rating_to_winner", "rating_to_loser",
                   "round", "category", "factor"]
        details.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=columns)

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def publish_championship_details_sheet(tournaments: models.Tournaments, rankings: models.Rankings,
                                       players: models.Players, tid: str, prev_tid: str, upload: bool) -> None:
    """Format and publish championship details of given tournament into sheets"""

    sheet_name = tournaments[tid]["sheet_name"].iloc[0]
    sheet_name = sheet_name.replace(cfg["sheetname"]["tournaments_key"], cfg["sheetname"]["championship_details_key"])

    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)

    championship_details = rankings.get_championship_details(tid)

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        headers = [cfg["labels"][key] for key in ["Player", "Category", "Best Round", "Championship Points"]]
        columns = ["name", "category", "best_round", "points"]
        championship_details.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=columns)

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def publish_statistics_sheet(tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                             tid: str, prev_tid: str, upload: bool = False) -> None:
    """ Copy details from log and output details of given tournament"""
    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)
    sheet_name = cfg["sheetname"]["statistics_key"]

    stats = rankings.get_statistics()
    headers = models.categories + ['total'] + models.categories + ['total']

    with _get_writer(xlsx_filename, sheet_name) as writer:
        stats.to_excel(writer, sheet_name=sheet_name, index=True, header=headers, index_label=cfg["labels"]["tid"])

        ws = writer.book[sheet_name]
        ws.insert_rows(0, 1)

        n_headers = len(headers)
        starting_cell, ending_cell = rowcol_to_a1(row=1, col=2), rowcol_to_a1(row=1, col=n_headers / 2 + 1)
        ws[starting_cell] = cfg["labels"]["Cumulated"]
        ws.merge_cells(f'{starting_cell}:{ending_cell}')

        starting_cell, ending_cell = rowcol_to_a1(row=1, col=n_headers / 2 + 2), rowcol_to_a1(row=1, col=n_headers + 1)
        ws[starting_cell] = cfg["labels"]["By Tournament"]
        ws.merge_cells(f'{starting_cell}:{ending_cell}')

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def publish_championship_sheets(tournaments: models.Tournaments, rankings: models.Rankings, players: models.Players,
                                tid: str, prev_tid: str, upload: bool = False) -> None:
    """Publish championship sheets, per category"""
    xlsx_filename = cfg["io"]["data_folder"] + cfg["io"]["publish_filename"].replace("NN", tid)

    # Rankings to be sorted by cum_points
    this_ranking = rankings[tid]
    prev_ranking = rankings[prev_tid]

    # Format data and columns to write into the file
    this_ranking = this_ranking.merge(players.players_df.loc[:, ["name", "city", "affiliation"]], on="pid")

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4", "E4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    headers = [cfg["labels"][key] for key in ["Position", "Championship Points", "Participations",
                                              "Player", "City", "Association", "Selected Tournaments"]]
    columns = ["position", "formatted points", "participations", "name", "city", "affiliation", "selected_tids"]

    for cat, point_col, selected_tids_col, participations_col in zip(models.categories,
                                                                     rankings.cum_points_cat_columns(),
                                                                     rankings.cum_tids_cat_columns(),
                                                                     rankings.participations_cat_columns()):
        sheet_name = tournaments[tid]["sheet_name"].iloc[0]
        sheet_name = sheet_name.replace(cfg["sheetname"]["tournaments_key"], cat.title())

        # Filter inactive players or players that didn't played any tournament
        unsorted_ranking = this_ranking.loc[this_ranking.loc[:, point_col] > 0].copy()
        sorted_ranking = unsorted_ranking.sort_values([point_col, participations_col], ascending=[False, True])

        if sorted_ranking.empty:
            continue

        # Format data and columns to write into the file
        sorted_ranking.insert(0, "position", range(1, len(sorted_ranking.index)+1))
        criteria = sorted_ranking[point_col] == sorted_ranking[point_col].shift(1)
        criteria &= sorted_ranking[participations_col] == sorted_ranking[participations_col].shift(1)
        sorted_ranking.loc[criteria, "position"] = None  # Equivalent positions will be deleted to avoid misleading

        sorted_ranking = sorted_ranking.merge(prev_ranking.loc[:, ["pid", point_col]], on="pid", suffixes=("", "_prev"))
        sorted_ranking.insert(6, "formatted points", sorted_ranking.apply(
            lambda row: _format_diff(row[point_col], row[point_col + "_prev"]), axis="columns"))
        sorted_ranking.loc[:, selected_tids_col] = sorted_ranking.loc[:, selected_tids_col].str.replace(tid[:-3], "")

        replacements = {"participations": participations_col, "selected_tids": selected_tids_col}
        _columns = [replacements[c] if c in replacements else c for c in columns]

        with _get_writer(xlsx_filename, sheet_name) as writer:
            sorted_ranking.to_excel(writer, sheet_name=sheet_name, index=False, header=headers, columns=_columns)

            # publish and format tournament metadata
            ws = writer.book[sheet_name]
            _publish_tournament_metadata(ws, tournaments[tid])
            _bold_and_center(ws, to_bold, to_center)  # FIXME This should be part of publish metadata, to merge the right cells

        if upload:
            load_and_upload_sheet(xlsx_filename, sheet_name, cfg["io"]["temporal_spreadsheet_id"])


def in_colab() -> bool:
    # Verify if it is running on colab
    try:
        import google.colab
        _in_colab = True
    except:
        _in_colab = False

    return _in_colab


def _get_gc() -> gspread.Client:
    try:
        if in_colab():
            from google.colab import auth
            auth.authenticate_user()
            from oauth2client.client import GoogleCredentials  # type: ignore
            gc = gspread.authorize(GoogleCredentials.get_application_default())
        else:
            gc = gspread.oauth()
    except FileNotFoundError:
        print("The .json key file has not been configured. Upload will fail.")
        raise ConnectionError
    # except OSError:
    #     print("Connection failure. Upload will fail.")

    return gc


def load_and_upload_sheet(filename: str, sheet_name: str, spreadsheet_id: str) -> None:
    df = pd.read_excel(filename, sheet_name, index_col=None, header=None, na_filter=False)
    upload_sheet_from_df(spreadsheet_id, sheet_name, df, include_df_headers=False)  # index is pid


def create_n_tour_sheet(spreadsheet_id: str, tid: str) -> None:
    """
    Create sheet corresponding to n_tour tournament by duplication of the first-tournament sheet.
    A new sheeet is created in given spreadsheet_id as follows:
    1- first-tournament sheet is duplicated
    2- Two replacements are performed in the new sheet, considering n_tour.
       For example, if n_tour=4, value of A1 cell and sheet title will change 'Tournament 01'->'Tournament 04'
    :param spreadsheet_id: spreadsheet where duplication will be performed
    :param tid: tournament to create
    :return: None
    """
    first_key = f'{cfg["labels"]["Tournament"]} 01'
    replacement_key = f'{cfg["labels"]["Tournament"]} {tid[-2:]}'

    try:
        gc = _get_gc()
        wb = gc.open_by_key(spreadsheet_id)
        sheetname_listed = [ws.title for ws in wb.worksheets() if first_key in ws.title]
        if sheetname_listed:
            sheetname = sheetname_listed[0]
            new_sheetname = sheetname.replace(first_key, replacement_key)
            if new_sheetname in [ws.title for ws in wb.worksheets()]:
                out_ws = wb.worksheet(new_sheetname)
                wb.del_worksheet(out_ws)
            ws = wb.worksheet(sheetname)
            dup_ws = wb.duplicate_sheet(ws.id, new_sheet_name=new_sheetname)
            dup_cell_value = dup_ws.acell('A1', value_render_option='FORMULA').value
            dup_ws.update_acell('A1', dup_cell_value.replace(first_key, replacement_key))
            print("<<<Creating", new_sheetname, "from", sheetname, "in", spreadsheet_id, sep="\t")
        else:
            print("FAILED TO DUPLICATE", first_key, "do not exist in", spreadsheet_id, sep="\t")

    except ConnectionError:
        print("<<<Connection Error. FAILED TO DUPLICATE", first_key, "in", spreadsheet_id, sep="\t")


def publish_to_web(tid: str, show_on_web=False) -> None:
    if show_on_web:
        for spreadsheet_id in cfg["io"]["published_on_web_spreadsheets_id"]:
            create_n_tour_sheet(spreadsheet_id, tid)


def load_temp_players_ranking() -> Tuple[models.Players, models.Rankings]:
    """returns players_temp, ranking_temp"""
    # Loading temp ranking and players. It should be deleted after a successful preprocessing
    players_temp_file = os.path.join(cfg["io"]["data_folder"], cfg["io"]["players_temp_file"])
    ranking_temp_file = os.path.join(cfg["io"]["data_folder"], cfg["io"]["ranking_temp_file"])

    if os.path.exists(players_temp_file):
        players_temp = load_from_pickle(cfg["io"]["players_temp_file"])
        print("Resume preprocessing...")
    else:
        players_temp = models.Players()

    if os.path.exists(ranking_temp_file):
        ranking_temp = load_from_pickle(cfg["io"]["ranking_temp_file"])
        print("Resume preprocessing...")
    else:
        ranking_temp = models.Rankings()

    return players_temp, ranking_temp


def save_temp_players_ranking(players_temp: models.Players, ranking_temp: models.Rankings) -> None:
    """returns players_temp, ranking_temp"""
    # Loading temp ranking and players. It shuould be deleted after a successful preprocessing
    players_temp_file = os.path.join(cfg["io"]["data_folder"], cfg["io"]["players_temp_file"])
    ranking_temp_file = os.path.join(cfg["io"]["data_folder"], cfg["io"]["ranking_temp_file"])
    print("<Saving\t\tTemps to resume preprocessing (if necessary)", ranking_temp_file, players_temp_file, "\n")
    with open(players_temp_file, 'wb') as ptf, open(ranking_temp_file, 'wb') as rtf:
        # Pickle the 'data' dictionary using the highest protocol available.
        pickle.dump(players_temp, ptf, pickle.HIGHEST_PROTOCOL)
        pickle.dump(ranking_temp, rtf, pickle.HIGHEST_PROTOCOL)


def remove_temp_players_ranking() -> None:
    players_temp_file = os.path.join(cfg["io"]["data_folder"], cfg["io"]["players_temp_file"])
    ranking_temp_file = os.path.join(cfg["io"]["data_folder"], cfg["io"]["ranking_temp_file"])
    print("Removing temp files created to resume preprocessing", players_temp_file, ranking_temp_file)
    if os.path.exists(players_temp_file):
        os.remove(players_temp_file)
    if os.path.exists(ranking_temp_file):
        os.remove(ranking_temp_file)


def save_to_pickle(players: models.Players = None, rankings: models.Rankings = None,
                   tournaments: models.Tournaments = None) -> None:
    objects = [players, rankings, tournaments]
    filenames = [cfg["io"]["players_pickle"], cfg["io"]["rankings_pickle"], cfg["io"]["tournaments_pickle"]]
    objects_filenames = [(obj, fn) for obj, fn in zip(objects, filenames) if obj]

    for obj, fn in objects_filenames:
        print('<<<Saving', fn, 'in', cfg["io"]["data_folder"], sep="\t")
        with open(os.path.join(cfg["io"]["data_folder"], fn), 'wb') as fo:
            pickle.dump(obj, fo, pickle.HIGHEST_PROTOCOL)


def load_from_pickle(filename: str) -> Any:
    print('>>>Loading', filename, 'from', cfg["io"]["data_folder"], sep="\t")
    with open(os.path.join(cfg["io"]["data_folder"], filename), 'rb') as fo:
        obj = pickle.load(fo)

    return obj
