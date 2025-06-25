from typing import List

import pandas as pd
from gspread.utils import rowcol_to_a1
from openpyxl.styles import Alignment, Font

from ranking_table_tennis import models
from ranking_table_tennis.configs import ConfigManager
from ranking_table_tennis.helpers.excel import _get_writer
from ranking_table_tennis.helpers.gspread import load_and_upload_sheet
from ranking_table_tennis.helpers.markdown import (
    publish_sheet_as_markdown,
    publish_stat_plot,
    publish_tournament_metadata_as_markdown,
)


def publish_championship_details_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    prev_tid: str,
    upload: bool,
) -> None:
    """Format and publish championship details of given tournament into sheets"""
    cfg = ConfigManager().current_config

    sheet_name = tournaments[tid]["sheet_name"].iloc[0]
    sheet_name = sheet_name.replace(
        cfg.sheetname.tournaments_key, cfg.sheetname.championship_details_key
    )

    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)

    to_bold = ["A1", "A2", "A3", "A4", "B4", "C4", "D4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    headers = [
        cfg.labels[key] for key in ["Player", "Category", "Best_Round", "Championship_Points"]
    ]
    columns = ["name", "category", "best_round", "points"]

    championship_details = (
        rankings.get_championship_details(tid)
        .loc[:, columns]
        .sort_values(
            ["category", "points", "name"], ascending=[True, False, True], ignore_index=True
        )
        .pipe(_insert_empty_row_between_categories)
    )

    with _get_writer(xlsx_filename, sheet_name) as writer:
        championship_details.to_excel(
            writer, sheet_name=sheet_name, index=False, header=headers, columns=columns
        )

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

    sheet_name_for_md = cfg.sheetname.championship_details_key
    publish_sheet_as_markdown(championship_details[columns], headers, sheet_name_for_md, tid)


def publish_statistics_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    prev_tid: str,
    upload: bool = False,
) -> None:
    """Copy details from log and output details of given tournament"""
    cfg = ConfigManager().current_config
    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)
    sheet_name = cfg.sheetname.statistics_key

    stats = rankings.get_statistics(tid)
    headers = cfg.categories + ["total"] + cfg.categories + ["total"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        stats.to_excel(
            writer,
            sheet_name=sheet_name,
            index=True,
            header=headers,
            index_label=cfg.labels.tid,
        )

        ws = writer.book[sheet_name]
        ws.insert_rows(0, 1)

        n_headers = len(headers)
        starting_cell, ending_cell = rowcol_to_a1(row=1, col=2), rowcol_to_a1(
            row=1, col=n_headers / 2 + 1
        )
        ws[starting_cell] = cfg.labels.Cumulated
        ws.merge_cells(f"{starting_cell}:{ending_cell}")

        starting_cell, ending_cell = rowcol_to_a1(row=1, col=n_headers / 2 + 2), rowcol_to_a1(
            row=1, col=n_headers + 1
        )
        ws[starting_cell] = cfg.labels.By_Tournament
        ws.merge_cells(f"{starting_cell}:{ending_cell}")

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

    sheet_name_for_md = f"{cfg.sheetname.statistics_key} {cfg.labels.Cumulated}"
    publish_sheet_as_markdown(
        stats.iloc[:, : stats.shape[1] // 2], headers, sheet_name_for_md, tid, index=True
    )
    publish_stat_plot(
        stats.iloc[:, : stats.shape[1] // 2 - 2],
        headers[: stats.shape[1] // 2 - 2],
        tid,
        sheet_name_for_md,
    )

    sheet_name_for_md = f"{cfg.sheetname.statistics_key} {cfg.labels.By_Tournament}"
    publish_sheet_as_markdown(
        stats.iloc[:, stats.shape[1] // 2 :], headers, sheet_name_for_md, tid, index=True
    )
    publish_stat_plot(
        stats.iloc[:, stats.shape[1] // 2 : -2],
        headers[stats.shape[1] // 2 : -2],
        tid,
        sheet_name_for_md,
    )


def publish_championship_sheets(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    prev_tid: str,
    upload: bool = False,
) -> None:
    """Publish championship sheets, per category"""
    cfg = ConfigManager().current_config
    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)

    # Rankings to be sorted by cum_points
    this_ranking = rankings[tid]
    prev_ranking = rankings[prev_tid]

    # Format data and columns to write into the file
    this_ranking = this_ranking.merge(
        players.players_df.loc[:, ["name", "city", "affiliation"]], on="pid"
    )

    to_bold = ["A1", "A2", "A3", "A4", "B4", "C4", "D4", "E4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    headers = [
        cfg.labels[key]
        for key in [
            "Position",
            "Championship_Points",
            "Participations",
            "Player",
            "City",
            "Association",
            "Selected_Tournaments",
        ]
    ]
    columns = [
        "position",
        "formatted points",
        "participations",
        "name",
        "city",
        "affiliation",
        "selected_tids",
    ]

    for cat, point_col, selected_tids_col, participations_col in zip(
        cfg.categories,
        rankings.cum_points_cat_columns(),
        rankings.cum_tids_cat_columns(),
        rankings.participations_cat_columns(),
    ):
        sheet_name = tournaments[tid]["sheet_name"].iloc[0]
        sheet_name = sheet_name.replace(cfg.sheetname.tournaments_key, cat.title())

        # Filter inactive players or players that didn't played any tournament
        unsorted_ranking = this_ranking.loc[this_ranking.loc[:, point_col] > 0].copy()
        sorted_ranking = unsorted_ranking.sort_values(
            [point_col, participations_col, "name"], ascending=[False, True, True]
        )

        if sorted_ranking.empty:
            continue

        # Format data and columns to write into the file
        sorted_ranking.insert(0, "position", range(1, len(sorted_ranking.index) + 1))
        criteria = sorted_ranking[point_col] == sorted_ranking[point_col].shift(1)
        criteria &= sorted_ranking[participations_col] == sorted_ranking[participations_col].shift(
            1
        )
        sorted_ranking.loc[criteria, "position"] = (
            None  # Equivalent positions will be deleted to avoid misleading
        )

        sorted_ranking = sorted_ranking.merge(
            prev_ranking.loc[:, ["pid", point_col]], on="pid", suffixes=("", "_prev")
        )
        sorted_ranking.insert(
            6,
            "formatted points",
            sorted_ranking.apply(
                lambda row: _format_diff(row[point_col], row[point_col + "_prev"]), axis="columns"
            ),
        )
        sorted_ranking.loc[:, selected_tids_col] = sorted_ranking.loc[
            :, selected_tids_col
        ].str.replace(tid[:-3], "")

        replacements = {"participations": participations_col, "selected_tids": selected_tids_col}
        _columns = [replacements[c] if c in replacements else c for c in columns]

        with _get_writer(xlsx_filename, sheet_name) as writer:
            sorted_ranking.to_excel(
                writer, sheet_name=sheet_name, index=False, header=headers, columns=_columns
            )

            # publish and format tournament metadata
            ws = writer.book[sheet_name]
            _publish_tournament_metadata(ws, tournaments[tid])
            _bold_and_center(
                ws, to_bold, to_center
            )  # FIXME This should be part of publish metadata, to merge the right cells

        if upload:
            load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

        # Workaround to remove nans before converting to as markdown
        sorted_rankind_md = (
            sorted_ranking.fillna({_columns[0]: "0"})
            .astype({_columns[0]: "int"})
            .replace({_columns[0]: 0}, "")
        )
        sheet_name_for_md = cat.title()
        publish_sheet_as_markdown(sorted_rankind_md[_columns], headers, sheet_name_for_md, tid)


def publish_rating_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    prev_tid: str,
    upload=False,
    all_players: bool = False,
) -> None:
    """Format a ranking to be published into a rating sheet"""
    cfg = ConfigManager().current_config
    sheet_name = tournaments[tid]["sheet_name"].iloc[0]
    sheet_name = sheet_name.replace(cfg.sheetname.tournaments_key, cfg.labels.Rating)
    if all_players:
        sheet_name += "_a"

    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)

    # Rankings sorted by rating
    this_ranking_df = rankings[tid].copy()
    prev_ranking_df = rankings[prev_tid].copy()

    # Filter inactive players or players that didn't played any tournament
    if not all_players:
        nonzero_points = this_ranking_df.loc[:, rankings.cum_points_cat_columns()].any(
            axis="columns"
        )
        this_ranking_df = this_ranking_df.loc[this_ranking_df.active | nonzero_points]
    # # FIXME there must be a special treatment for fan category

    # Format data and columns to write into the file
    this_ranking_df = this_ranking_df.merge(
        players.players_df.loc[:, ["name", "city", "affiliation"]], on="pid"
    )
    this_ranking_df = this_ranking_df.merge(
        prev_ranking_df.loc[:, ["pid", "rating"]], on="pid", suffixes=("", "_prev")
    )
    this_ranking_df.insert(
        6,
        "formatted rating",
        this_ranking_df.apply(
            lambda row: _format_diff(row["rating"], row["rating_prev"]), axis="columns"
        ),
    )

    to_bold = ["A1", "A2", "A3", "A4", "B4", "C4", "D4", "E4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    headers = [cfg.labels[key] for key in ["Category", "Rating", "Player", "City", "Association"]]
    columns = ["category", "formatted rating", "name", "city", "affiliation"]

    if all_players:
        this_ranking_df = this_ranking_df.sort_values(["name"], ascending=[True]).loc[:, columns]
    else:
        this_ranking_df = (
            this_ranking_df.sort_values(["rating", "name"], ascending=[False, True])
            .loc[:, columns]
            .pipe(_insert_empty_row_between_categories)
        )

    with _get_writer(xlsx_filename, sheet_name) as writer:
        this_ranking_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=headers, columns=columns
        )

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

    sheet_name_for_md = cfg.labels.Rating
    if all_players:
        sheet_name_for_md += "_Ampliado"
    publish_sheet_as_markdown(this_ranking_df[columns], headers, sheet_name_for_md, tid)
    publish_tournament_metadata_as_markdown(tid, tournaments[tid])


def publish_initial_rating_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    upload=False,
) -> None:
    """Format a ranking to be published into a rating sheet"""
    cfg = ConfigManager().current_config
    sheet_name = cfg.sheetname.initial_ranking
    sheet_name = sheet_name.replace(cfg.sheetname.rankings_key, cfg.labels.Rating)

    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)

    # Rankings sorted by rating and name
    initial_tid = cfg.initial_metadata.initial_tid
    this_ranking_df = (
        rankings[initial_tid]
        # Format data and columns to write into the file
        .merge(players.players_df.loc[:, ["name", "city", "affiliation"]], on="pid").sort_values(
            ["rating", "name"], ascending=[False, True]
        )
    )

    to_bold = ["A1", "B1", "C1", "D1"]
    to_center = to_bold

    headers = [cfg.labels[key] for key in ["Rating", "Player", "City", "Association"]]
    columns = ["rating", "name", "city", "affiliation"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        this_ranking_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=headers, columns=columns
        )

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

    # Adds a label for initial rating so it does not overwrite current rating file
    sheet_name_for_md = f"{cfg.labels.Rating}_{cfg.initial_metadata.tournament_name.split()[0]}"
    publish_sheet_as_markdown(this_ranking_df[columns], headers, sheet_name_for_md, tid)


def publish_histories_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    prev_tid: str,
    upload=False,
) -> None:
    """Format histories to be published into a sheet"""
    cfg = ConfigManager().current_config
    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)
    sheet_name = cfg.sheetname.histories

    tids_to_consider = players.history_df.tid <= tid
    history_df = players.history_df.loc[tids_to_consider].copy()

    # Match pid to get player's metadata into new columns of history_df
    history_df = pd.merge(history_df, players.players_df, on="pid")
    # Sort histories by name, category and tid
    history_df = history_df.sort_values(["name", "category", "tid"], ascending=[True, False, True])
    # Remove repeated strings to show a cleaner sheet
    history_df.loc[history_df["name"] == history_df["name"].shift(1), "name"] = ""
    # insert an empty row into history. This is a format workaround
    history_df = history_df.groupby("pid", sort=False, group_keys=False)[history_df.columns].apply(
        _keep_name_new_row
    )
    # Remove repeated strings to show a cleaner sheet
    history_df.loc[history_df["category"] == history_df["category"].shift(1), "category"] = ""

    headers = [cfg.labels[key] for key in ["Player", "Category", "Best_Round", "Tournament"]]
    columns = ["name", "category", "best_round", "tid"]

    with _get_writer(xlsx_filename, sheet_name) as writer:
        history_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=headers, columns=columns
        )

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

    sheet_name_for_md = cfg.sheetname.histories
    publish_sheet_as_markdown(history_df[columns], headers, sheet_name_for_md, tid)


def publish_rating_details_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    prev_tid: str,
    upload,
) -> None:
    """Format and publish rating details of given tournament into a sheet"""
    cfg = ConfigManager().current_config
    sheet_name = tournaments[tid]["sheet_name"].iloc[0]
    sheet_name = sheet_name.replace(cfg.sheetname.tournaments_key, cfg.sheetname.rating_details_key)

    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)

    to_bold = ["A1", "A2", "A3", "A4", "B4", "C4", "D4", "E4", "F4", "G4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    headers = [
        cfg.labels[key]
        for key in [
            "Winner",
            "Loser",
            "Difference",
            "Winner_Points",
            "Loser_Points",
            "Round",
            "Category",
            "Factor",
        ]
    ]
    columns = [
        "winner_name_rating",
        "loser_name_rating",
        "diff_rating",
        "rating_to_winner",
        "rating_to_loser",
        "round",
        "category",
        "factor",
    ]

    details = (
        rankings.get_rating_details(tid)
        .assign(
            winner_name_rating=lambda df: df["winner"]
            + " ("
            + df["winner_rating"].astype(int).astype(str)
            + ")",
            loser_name_rating=lambda df: df["loser"]
            + " ("
            + df["loser_rating"].astype(int).astype(str)
            + ")",
            diff_rating=lambda df: (df["winner_rating"] - df["loser_rating"])
            .astype(int)
            .astype(str),
            factor=lambda df: (df["factor"] / cfg.compute.rating_factor)
            .astype(str)
            .str.replace(".0", "", regex=False),
            rating_to_winner=lambda df: df["rating_to_winner"].astype(int).astype(str),
            rating_to_loser=lambda df: df["rating_to_loser"].astype(int).astype(str),
        )
        .sort_values(
            ["category", "round", "winner_name_rating", "loser_name_rating"],
            ascending=[True, True, True, True],
        )
        .loc[:, columns]
        .pipe(_insert_empty_row_between_categories)
    )

    with _get_writer(xlsx_filename, sheet_name) as writer:
        details.to_excel(
            writer, sheet_name=sheet_name, index=False, header=headers, columns=columns
        )

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    if upload:
        load_and_upload_sheet(xlsx_filename, sheet_name, cfg.io.temporal_spreadsheet_id)

    sheet_name_for_md = cfg.sheetname.rating_details_key
    publish_sheet_as_markdown(details[columns], headers, sheet_name_for_md, tid)


def publish_matches_sheet(
    tournaments: models.Tournaments,
    rankings: models.Rankings,
    players: models.Players,
    tid: str,
    upload,
) -> None:
    """Format and publish matches of given tournament into a sheet"""
    cfg = ConfigManager().current_config

    sheet_name = tournaments[tid]["sheet_name"].iloc[0]

    xlsx_filename = cfg.io.data_folder + cfg.io.xlsx.publish_filename.replace("NN", tid)

    to_bold = ["A1", "A2", "A3", "A4", "B4", "C4", "D4", "E4", "F4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    headers = [
        cfg.labels.Player + " A",
        cfg.labels.Player + " B",
        "Sets A",
        "Sets B",
        cfg.labels.Round,
        cfg.labels.Category,
    ]
    columns = [
        "player_a",
        "player_b",
        "sets_a",
        "sets_b",
        "round",
        "category",
    ]

    matches = (
        tournaments.get_matches(tid, False, [])
        .loc[:, columns]
        .sort_values(["category", "round", "player_a", "player_b"], ascending=True)
        .pipe(_insert_empty_row_between_categories)
    )

    with _get_writer(xlsx_filename, sheet_name) as writer:
        matches.to_excel(
            writer, sheet_name=sheet_name, index=False, header=headers, columns=columns
        )

        # publish and format tournament metadata
        ws = writer.book[sheet_name]
        _publish_tournament_metadata(ws, tournaments[tid])
        _bold_and_center(ws, to_bold, to_center)

    sheet_name_for_md = cfg.sheetname.tournaments_key
    publish_sheet_as_markdown(matches[columns], headers, sheet_name_for_md, tid)


def _publish_tournament_metadata(ws, tournament_tid: pd.DataFrame) -> None:
    cfg = ConfigManager().current_config
    ws.insert_rows(0, 3)
    ws["A1"] = cfg.labels.Tournament_name
    ws["B1"] = tournament_tid["tournament_name"].iloc[0]
    ws.merge_cells("B1:E1")
    ws["A2"] = cfg.labels.Date
    ws["B2"] = tournament_tid["date"].iloc[0].strftime("%Y %m %d")
    ws.merge_cells("B2:E2")
    ws["A3"] = cfg.labels.Location
    ws["B3"] = tournament_tid["location"].iloc[0]
    ws.merge_cells("B3:E3")


def _bold_and_center(ws, to_bold: List[str], to_center: List[str]) -> None:
    """Bold and center cells in given worksheet (ws)"""
    for colrow in to_bold:
        cell = ws[colrow]
        cell.font = Font(bold=True)
    for colrow in to_center:
        cell = ws[colrow]
        cell.alignment = Alignment(horizontal="center")


def _format_diff(new_value: float, prev_value: float) -> str:
    """Return a formated str that indicates +##, -## or ="""
    diff = new_value - prev_value
    diff_str = f"{diff:.0f}"
    if diff > 0:
        diff_str = "+" + diff_str
    elif diff == 0:
        diff_str = "="
    formatted_str = f"{new_value:.0f} ({diff_str})"

    return formatted_str


def _keep_name_new_row(df: pd.DataFrame) -> pd.DataFrame:
    """Function to insert row in the dataframe"""
    empty_row = pd.DataFrame(
        {
            "tid": "",
            "pid": "",
            "category": "",
            "best_round": "",
            "name": df.loc[df.first_valid_index(), "name"],
        },
        index=[-1],
    )
    df.loc[df.first_valid_index(), "name"] = ""

    return pd.concat([empty_row, df])


def _insert_empty_row_between_categories(df: pd.DataFrame) -> pd.DataFrame:
    """Insert empty row between categories in the dataframe"""
    # First row is ommited because is an empty row
    return (
        df.groupby("category", sort=False, as_index=False)[df.columns]
        .apply(_insert_empty_row)
        .iloc[1:]
    )


def _insert_empty_row(df: pd.DataFrame) -> pd.DataFrame:
    """Function to insert empty row in the dataframe"""
    empty_row = df.iloc[:1].copy().astype(str)
    empty_row.iloc[0] = ""

    return pd.concat([empty_row, df], ignore_index=True)
