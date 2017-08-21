import csv
import models
import os
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

import gspread
from oauth2client.service_account import ServiceAccountCredentials

__author__ = 'sebastian'

# Loads some names from config.yaml
with open("config.yaml", 'r') as cfgyaml:
    try:
        cfg = yaml.load(cfgyaml)
    except yaml.YAMLError as exc:
        print(exc)

# Drive authorization
scope = ['https://spreadsheets.google.com/feeds']
key_filename = "key-for-gspread.json"
credentials = ServiceAccountCredentials.from_json_keyfile_name(key_filename, scope)
gc = gspread.authorize(credentials)


def get_sheetnames_by_date(filter_key=""):
    tournaments_xlsx = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]
    wb = load_workbook(tournaments_xlsx, read_only=True)
    sheetnames = [s for s in wb.sheetnames if filter_key in s]
    namesdates = [(name, load_tournament_xlsx(name).date) for name in sheetnames]
    namesdates.sort(key=lambda p: p[1])

    return [name for name, date in namesdates]


def load_sheet_workbook(filename, sheetname, first_row=1):
    print(">Reading\t", sheetname, "\tfrom\t", filename)
    wb = load_workbook(filename, read_only=True)
    ws = wb.get_sheet_by_name(sheetname)

    ws.calculate_dimension(force=True)

    list_to_return = []
    for row in ws.rows:
        aux_row = []
        empty_row = True
        for cell in row:
            if cell.value is None:
                aux_row.append("")
            else:
                empty_row = False
                aux_row.append(cell.value)
        if not empty_row:
            list_to_return.append(aux_row[:ws.max_column])
    return list_to_return[first_row:]


def _wb_ws_to_save(filename, sheetname, overwrite):
    print("<<<Saving\t", sheetname, "\tin\t", filename)
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        if overwrite and sheetname in wb:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        if sheetname in wb:
            ws = wb.get_sheet_by_name(sheetname)
        else:
            ws = wb.create_sheet()
    else:
        wb = Workbook()
        ws = wb.active

    ws.title = sheetname

    return wb, ws


def _bold_and_center(ws, to_bold, to_center):
    """Bold and center cells in given worksheet (ws)"""
    for colrow in to_bold:
        cell = ws[colrow]
        cell.font = Font(bold=True)
    for colrow in to_center:
        cell = ws[colrow]
        cell.alignment = Alignment(horizontal='center')


def save_sheet_workbook(filename, sheetname, headers, list_to_save, overwrite=True):
    wb, ws = _wb_ws_to_save(filename, sheetname, overwrite)

    ws.append(headers)
    
    for row in list_to_save:
        ws.append(row)

    for col in range(1, ws.max_column+1):
        cell = ws.cell(column=col, row=1)
        cell.font = Font(bold=True)

    wb.save(filename)


def _format_ranking_header_and_list(ranking, players):
    headers = [cfg["labels"][key] for key in ["PID", "Rating Points", "Bonus Points",
                                              "Active Player", "Category", "Player"]]
    list_to_save = [[e.pid, e.rating, e.bonus, cfg["activeplayer"][e.active],
                     e.category, players[e.pid].name] for e in ranking]
    list_to_save.sort(key=lambda l: l[1], reverse=True)

    return headers, list_to_save


def save_ranking_sheet(filename, sheetname, ranking, players, overwrite=True, replace_key=True):
    if replace_key:
        sheetname = sheetname.replace(cfg["sheetname"]["tournaments_key"], cfg["sheetname"]["rankings_key"])

    wb, ws = _wb_ws_to_save(filename, sheetname, overwrite)
    headers, list_to_save = _format_ranking_header_and_list(ranking, players)

    ws["A1"] = cfg["labels"]["Tournament name"]
    ws["B1"] = ranking.tournament_name
    ws["A2"] = cfg["labels"]["Date"]
    ws["B2"] = ranking.date
    ws["A3"] = cfg["labels"]["Location"]
    ws["B3"] = ranking.location
    ws["A4"] = cfg["labels"]["tid"]
    ws["B4"] = ranking.tid

    ws.append(headers)

    for row in list_to_save:
        ws.append(row)

    wb.save(filename)


def load_ranking_sheet(filename, sheetname, replace_key=True):
    """Load a ranking from a xlsx sheet and return a Ranking object

    :param filename: ranking workbook
    :param sheetname: tournament sheetname (will be replaced with ranking sheetname by default)
    :param replace_key: Default True. It will replace tournament key with ranking key.
    If False, sheetname will be used as it is given.

    :return: Ranking object
    """
    if replace_key:
        sheetname = sheetname.replace(cfg["sheetname"]["tournaments_key"], cfg["sheetname"]["rankings_key"])

    # TODO check if date is being read properly
    raw_ranking = load_sheet_workbook(filename, sheetname, first_row=0)
    ranking = models.Ranking(raw_ranking[0][1], raw_ranking[1][1], raw_ranking[2][1], raw_ranking[3][1])
    ranking.load_list([[rr[0], rr[1], rr[2], rr[3] == cfg["activeplayer"][True], rr[4]] for rr in raw_ranking[5:]])

    return ranking


def load_tournament_csv(filename):
    """Load a tournament csv and return a Tournament object"""
    with open(filename, 'r') as incsv:
        reader = csv.reader(incsv)
        tournament_list = [row for row in reader]
        return load_tournament_list(tournament_list)


def load_tournament_xlsx(sheet_name):
    """Load a tournament xlsx sheet and return a Tournament object

    tournaments xlsx database is defined in config.yaml
    """
    tournaments_xlsx = cfg["io"]["data_folder"] + cfg["io"]["tournaments_filename"]

    return load_tournament_list(load_sheet_workbook(tournaments_xlsx, sheet_name, 0))


def load_tournament_list(tournament_list):
    """Load a tournament list sheet and return a Tournament object
    name = cell(B1)
    date = cell(B2)
    location = cell(B3)
    matches should be from sixth row containing:
    player1, player2, sets1, sets2, match_round, category
    """
    name = tournament_list[0][1]
    date = tournament_list[1][1]
    location = tournament_list[2][1]

    tournament = models.Tournament(name, date, location)

    # Reformated list of matches
    for player1, player2, sets1, sets2, round_match, category in tournament_list[5:]:
        # workaround to add extra bonus points from match list
        if int(sets1) >= 10 and int(sets2) >= 10:
            winner_name = cfg["aux"]["flag promotion"]
            loser_name = player2
        elif int(sets1) < 0 and int(sets2) < 0:
            winner_name = cfg["aux"]["flag add bonus"]
            loser_name = player2
        elif int(sets1) > int(sets2):
            winner_name = player1
            loser_name = player2
        elif int(sets1) < int(sets2):
            winner_name = player2
            loser_name = player1
        else:
            print("Failed to process matches, a tie was found between %s and %s" % (player1, player2))
            break
        tournament.add_match(winner_name, loser_name, round_match, category)

    return tournament


def _format_diff(diff):
    """Return a formated str that indicates +##, -## or ="""
    diff_str = str(diff)
    if diff > 0:
        diff_str = "+" + diff_str
    elif diff == 0:
        diff_str = "="
    return diff_str


def publish_rating_sheet(filename, sheetname, ranking, players, old_ranking, overwrite=True):
    """ Format a ranking to be published into a rating sheet
    """
    sheetname = sheetname.replace(cfg["sheetname"]["tournaments_key"], cfg["labels"]["Rating Points"])

    wb, ws = _wb_ws_to_save(filename, sheetname, overwrite)

    ws["A1"] = cfg["labels"]["Tournament name"]
    ws["B1"] = ranking.tournament_name
    ws.merge_cells('B1:F1')
    ws["A2"] = cfg["labels"]["Date"]
    ws["B2"] = ranking.date
    ws.merge_cells('B2:F2')
    ws["A3"] = cfg["labels"]["Location"]
    ws["B3"] = ranking.location
    ws.merge_cells('B3:F3')

    ws.append(cfg["labels"][key] for key in ["Category", "Rating Points", "Player", "City",
                                             "Association", "Active Player"])

    list_to_save = [[e.category, (e.rating, old_ranking[e.pid].rating, e.bonus), players[e.pid].name,
                     players[e.pid].city, players[e.pid].association, cfg["activeplayer"][e.active]] for e in ranking
                    if e.bonus > 0 or ranking.tid < 6]  # Exclude players that didn't played for a long time

    for row in list_to_save:
        # Do not publish ratings of fans category
        if row[0] == models.categories[-1]:
            # Bonus points are used for fans. Negative values keep fans category at the end
            row[1] = (row[1][2]-100000, -1)

    for row in sorted(list_to_save, key=lambda l: l[1][0], reverse=True):
        if row[1][0] < 0:
            row[1] = "NA"  # FIXME should read the value from config
        else:
            # Save difference with previous rating
            diff = row[1][0] - row[1][1]
            row[1] = "%d (%s)" % (row[1][0], _format_diff(diff))
        ws.append(row)

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4", "E4", "F4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    _bold_and_center(ws, to_bold, to_center)

    wb.save(filename)


def publish_championship_sheet(filename, sheetname, ranking, players, old_ranking, overwrite=True):
    """ Format a ranking to be published into a rating sheet"""
    sheetname = sheetname.replace(cfg["sheetname"]["tournaments_key"], cfg["sheetname"]["championship_key"])

    wb, ws = _wb_ws_to_save(filename, sheetname, overwrite)

    ws["A1"] = cfg["labels"]["Tournament name"]
    ws["B1"] = ranking.tournament_name
    ws.merge_cells('B1:G1')
    ws["A2"] = cfg["labels"]["Date"]
    ws["B2"] = ranking.date
    ws.merge_cells('B2:G2')
    ws["A3"] = cfg["labels"]["Location"]
    ws["B3"] = ranking.location
    ws.merge_cells('B3:G3')

    ws.append(cfg["labels"][key] for key in ["Position", "Bonus Points", "Player", "City",
                                             "Association", "Active Player", "Category"])

    list_to_save = [[(e.bonus, old_ranking[e.pid].bonus), players[e.pid].name, players[e.pid].city,
                     players[e.pid].association, cfg["activeplayer"][e.active], e.category] for e in ranking
                    if e.bonus > 0 or ranking.tid < 6]  # Exclude players that didn't played for a long time
    sorted_list = sorted(list_to_save, key=lambda l: l[0][0], reverse=True)

    for i, row in enumerate(sorted_list):
        # Save difference with previous bonus
        diff = row[0][0] - row[0][1]
        row[0] = "%d (%s)" % (row[0][0], _format_diff(diff))
        ws.append([i + 1] + row)

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4", "E4", "F4", "G4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    _bold_and_center(ws, to_bold, to_center)

    wb.save(filename)


def publish_histories_sheet(filename, sheetname, players, tournament_sheetnames, overwrite=True):
    """ Format histories to be published into a sheet"""
    wb, ws = _wb_ws_to_save(filename, sheetname, overwrite)

    histories = []
    for player in sorted(players, key=lambda l: l.name):
        if len(player.sorted_history) > 0:
            histories.append([player.name, "", "", ""])
            old_cat = ""
            for cat, tid, best_round in player.sorted_history:
                if cat == old_cat:
                    cat = ""
                else:
                    old_cat = cat
                histories.append(["", cat, best_round, " ".join(tournament_sheetnames[tid - 1].split()[1:])])

    to_bold = ["A1", "A2", "A3", "A4"]
    to_center = to_bold

    _bold_and_center(ws, to_bold, to_center)

    save_sheet_workbook(filename, sheetname,
                        [cfg["labels"][key] for key in ["Player", "Category", "Best Round", "Tournament"]],
                        histories,
                        overwrite)


def _wb_ws_to_upload(spreadsheet_id, sheetname, num_rows, num_cols):
    print("<<<Saving\t", sheetname, "\tin\t", spreadsheet_id)
    wb = gc.open_by_key(spreadsheet_id)

    # Overwrites an existing sheet or creates a new one
    if sheetname in [ws.title for ws in wb.worksheets()]:
        ws = wb.worksheet(sheetname)
        ws.resize(rows=num_rows, cols=num_cols)
    else:
        ws = wb.add_worksheet(title=sheetname, rows=num_rows, cols=num_cols)

    return wb, ws


def upload_sheet(spreadsheet_id, sheetname, headers, rows_to_save):
    """ Saves headers and rows_to_save into given sheet_name.
        If sheet_name does not exist, it will be created. """
    num_rows = len(rows_to_save) + 1  # +1 because of header
    num_cols = len(headers)

    wb, ws = _wb_ws_to_upload(spreadsheet_id, sheetname, num_rows, num_cols)

    # Concatenation of all cells values to be updated in batch mode
    cell_list = ws.range("A1:" + ws.get_addr_int(row=num_rows, col=num_cols))
    for i, value in enumerate(headers + [v for row in rows_to_save for v in row]):
        cell_list[i].value = value

    ws.update_cells(cell_list)


def upload_ranking_sheet(spreadsheet_id, sheetname, ranking, players, replace_key=True):
    """ Saves ranking into given sheet_name.
        If sheet_name does not exist, it will be created. """
    if replace_key:
        sheetname = sheetname.replace(cfg["sheetname"]["tournaments_key"], cfg["sheetname"]["rankings_key"])

    headers, list_to_save = _format_ranking_header_and_list(ranking, players)

    num_cols = len(headers)
    num_rows = len(list_to_save) + 1 + 4  # +1 because of header + 4 because of tournament metadata

    wb, ws = _wb_ws_to_upload(spreadsheet_id, sheetname, num_rows, num_cols)

    ws.update_acell("A1", cfg["labels"]["Tournament name"])
    ws.update_acell("B1", ranking.tournament_name)
    ws.update_acell("A2", cfg["labels"]["Date"])
    ws.update_acell("B2", ranking.date)
    ws.update_acell("A3", cfg["labels"]["Location"])
    ws.update_acell("B3", ranking.location)
    ws.update_acell("A4", cfg["labels"]["tid"])
    ws.update_acell("B4", ranking.tid)

    # Concatenation of all cells values to be updated in batch mode
    cell_list = ws.range("A5:" + ws.get_addr_int(row=num_rows, col=num_cols))
    for i, value in enumerate(headers + [v for row in list_to_save for v in row]):
        cell_list[i].value = value

    ws.update_cells(cell_list)
